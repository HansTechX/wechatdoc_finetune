#!/usr/bin/env python3
"""
数据准备脚本 - 从 Excel 生成 Alpaca JSONL 格式微调数据
无参数时自动选择日期最新的 Excel 文件
"""

import os
import sys
import re
import glob
import json
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise ImportError("需要 openpyxl，请运行: pip install openpyxl")


class _Tee:
    """同时输出到控制台和日志文件"""
    def __init__(self, console, file):
        self.console = console
        self.file = file
    def write(self, text):
        self.console.write(text)
        self.file.write(text)
    def flush(self):
        self.console.flush()
        self.file.flush()


def find_latest_excel() -> str:
    """在项目根目录找日期最新的 Excel 文件"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = []
    for f in glob.glob(os.path.join(script_dir, "*.xlsx")):
        basename = os.path.basename(f)
        if basename.startswith("~$"):
            continue
        # 提取文件名中的日期 YYMMDD
        m = re.search(r"(\d{6})", basename)
        if m:
            candidates.append((m.group(1), f))
    if not candidates:
        raise FileNotFoundError("未找到 Excel 数据文件")
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def read_excel(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    data = {}
    for name in wb.sheetnames:
        ws = wb[name]
        data[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return data


def extract_system_prompt(rows: list) -> str:
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_system = header.index("System 提示词")
    col_train = header.index("是否训练")
    for row in rows[1:]:
        if len(row) > col_train and str(row[col_train]).strip() == "是":
            return str(row[col_system]).strip()
    return ""


def build_code_map(rows: list, prompt: str) -> tuple:
    """返回 (名称->编码映射, 小写索引, 编码->名称反查)"""
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_code = header.index("业务编码")
    col_name = header.index("业务名称")

    mapping = {}
    code_to_name = {}
    for row in rows[1:]:
        if not row or len(row) <= max(col_code, col_name):
            continue
        code, name = row[col_code], row[col_name]
        if code and name:
            code, name = str(code).strip(), str(name).strip()
            mapping[name] = code
            code_to_name[code] = name

    # 从提示词中补充
    for m in re.finditer(r'【(\d+)】(.+?)(?:[：:\n])', prompt):
        name, code = m.group(2).strip(), m.group(1)
        if name not in mapping:
            mapping[name] = code
            code_to_name.setdefault(code, name)

    # 大小写兼容
    lower_map = {k.lower(): v for k, v in mapping.items()}

    # 同义词
    synonyms = {"查附近网点": "查询服务网点"}
    for alias, target in synonyms.items():
        if alias not in mapping and target in mapping:
            mapping[alias] = mapping[target]
            lower_map[alias.lower()] = mapping[target]

    return mapping, lower_map, code_to_name


def convert_to_jsonl(rows: list, system_prompt: str, code_map: dict, lower_map: dict) -> tuple:
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_input = header.index("客户问题")
    col_output = header.index("人工标注结果")
    HEADER_VALUES = set(header)

    result = []
    skipped = 0
    unmapped_labels = {}
    code_dist = defaultdict(int)

    for row in rows[1:]:
        if not row or len(row) <= max(col_input, col_output):
            continue
        query, label = row[col_input], row[col_output]
        if not query or not label:
            continue
        query, label = str(query).strip(), str(label).strip()

        if query in HEADER_VALUES or label in HEADER_VALUES:
            skipped += 1
            continue

        code = code_map.get(label) or lower_map.get(label.lower())
        if not code:
            unmapped_labels[label] = unmapped_labels.get(label, 0) + 1
            skipped += 1
            continue

        result.append({"instruction": system_prompt, "input": query, "output": code})
        code_dist[code] += 1

    return result, skipped, unmapped_labels, code_dist


def save_jsonl(data: list, path: str):
    with open(path, "w", encoding="utf-8") as f:
        for item in data:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def print_distribution(title: str, code_dist: dict, code_to_name: dict, unmapped: dict, skipped: int):
    print(f"\n  [{title}]")
    for code in sorted(code_dist.keys(), key=int):
        name = code_to_name.get(code, "")
        print(f"    [{name}-{code}] : {code_dist[code]} 条")
    print(f"    合计: {sum(code_dist.values())} 条")
    if unmapped:
        print(f"    [!] 未映射标签 (跳过 {skipped} 条):")
        for label, count in unmapped.items():
            print(f"      - '{label}': {count} 条")


def update_dataset_info(dataset_name: str, train_file: str, val_file: str, dataset_dir: str):
    """注册数据集到 dataset_info.json"""
    info_path = os.path.join(dataset_dir, "dataset_info.json")
    if os.path.exists(info_path):
        with open(info_path, "r", encoding="utf-8") as f:
            info = json.load(f)
    else:
        info = {}
    info[dataset_name] = {
        "file_name": train_file,
        "columns": {
            "prompt": "instruction",
            "query": "input",
            "response": "output",
            "system": "instruction",
        },
    }
    if val_file:
        info[dataset_name]["val_file"] = val_file
    with open(info_path, "w", encoding="utf-8") as f:
        json.dump(info, f, ensure_ascii=False, indent=2)
    print(f"  已注册到 {os.path.join(dataset_dir, 'dataset_info.json')}")


def update_train_config(dataset_name: str):
    """自动更新 train_config.yaml 中的 dataset_name"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config", "train_config.yaml")
    if not os.path.exists(config_path):
        print(f"  [!] 未找到配置文件: {config_path}")
        return

    with open(config_path, "r", encoding="utf-8") as f:
        content = f.read()

    import re
    new_content, count = re.subn(
        r'(dataset_name:\s*").*?(")',
        rf'\g<1>{dataset_name}\g<2>',
        content,
    )
    if count > 0:
        with open(config_path, "w", encoding="utf-8") as f:
            f.write(new_content)
        print(f"  已更新 config/train_config.yaml: dataset_name -> {dataset_name}")
    else:
        print(f"  [!] 未在 train_config.yaml 中找到 dataset_name 字段")


def main():
    parser = argparse.ArgumentParser(description="从 Excel 生成微调训练数据")
    parser.add_argument("--input", nargs="?", default=None, help="Excel 文件路径 (默认选最新)")
    parser.add_argument("--output_dir", default="data", help="输出目录")
    parser.add_argument("--dataset_name", default="intent_cls", help="数据集名称")
    args = parser.parse_args()

    # 确定输入文件
    if args.input:
        excel_path = args.input
    else:
        excel_path = find_latest_excel()
        print(f"(自动选择最新文件)")

    print(f"数据文件: {os.path.basename(excel_path)}")
    os.makedirs(args.output_dir, exist_ok=True)

    # 日志输出到 outputs/prepare_data.log
    script_dir = os.path.dirname(os.path.abspath(__file__))
    log_path = os.path.join(script_dir, "outputs", "prepare_data.log")
    os.makedirs(os.path.dirname(log_path), exist_ok=True)
    _log_file = open(log_path, "a", encoding="utf-8")
    _log_file.write(f"\n{'='*60}\n[{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}]\n{'='*60}\n")
    sys.stdout = _Tee(sys.stdout, _log_file)

    sheets = read_excel(excel_path)

    # 提示词
    system_prompt = extract_system_prompt(sheets.get("提示词", []))
    print(f"提示词长度: {len(system_prompt)} 字符")

    # 编码映射
    code_map, lower_map, code_to_name = build_code_map(
        sheets.get("业务编码映射", []), system_prompt
    )
    print(f"编码映射: {len(code_map)} 条")

    # 训练集
    train_data, train_skip, train_unmapped, train_dist = convert_to_jsonl(
        sheets.get("训练集", []), system_prompt, code_map, lower_map
    )
    train_file = f"{args.dataset_name}.jsonl"
    train_path = os.path.join(args.output_dir, train_file)
    save_jsonl(train_data, train_path)
    print_distribution("训练集", train_dist, code_to_name, train_unmapped, train_skip)
    print(f"  已保存: {train_path}")

    # 验证集
    val_data, val_skip, val_unmapped, val_dist = convert_to_jsonl(
        sheets.get("验证集", []), system_prompt, code_map, lower_map
    )
    val_file = f"{args.dataset_name}_val.jsonl"
    val_path = os.path.join(args.output_dir, val_file)
    save_jsonl(val_data, val_path)
    print_distribution("验证集", val_dist, code_to_name, val_unmapped, val_skip)
    print(f"  已保存: {val_path}")

    # 注册 & 更新配置
    print()
    update_dataset_info(args.dataset_name, train_file, val_file, args.output_dir)
    update_train_config(args.dataset_name)

    print("\n数据准备完成。")

    sys.stdout = _log_file.console
    _log_file.close()


if __name__ == "__main__":
    main()
