#!/usr/bin/env python3
"""
数据准备脚本 - 从 Excel 生成 Alpaca JSONL 格式微调数据
无参数时自动选择日期最新的 Excel 文件
"""

import os
import sys
import re
import glob
import json
import logging
import argparse
from collections import defaultdict
from datetime import datetime

try:
    import openpyxl
except ImportError:
    raise ImportError("需要 openpyxl，请运行: pip install openpyxl")

logger = logging.getLogger("step1_prepare")


def setup_logger(log_path: str) -> logging.Logger:
    """配置日志：同时输出到控制台和文件，格式与 step2_train 一致"""
    log_dir = os.path.dirname(log_path)
    os.makedirs(log_dir, exist_ok=True)

    _logger = logging.getLogger("step1_prepare")
    _logger.setLevel(logging.DEBUG)

    fmt = logging.Formatter(
        "[%(asctime)s] [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)

    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)

    _logger.addHandler(ch)
    _logger.addHandler(fh)
    return _logger


def find_latest_excel() -> str:
    """在项目根目录找日期最新的 Excel 文件"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    candidates = []
    for f in glob.glob(os.path.join(script_dir, "*.xlsx")):
        basename = os.path.basename(f)
        if basename.startswith("~$"):
            continue
        # 提取文件名中的日期 YYMMDD
        m = re.search(r"(\d{6})", basename)
        if m:
            candidates.append((m.group(1), f))
    if not candidates:
        raise FileNotFoundError("未找到 Excel 数据文件")
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0][1]


def read_excel(excel_path: str) -> dict:
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    data = {}
    for name in wb.sheetnames:
        ws = wb[name]
        data[name] = [list(row) for row in ws.iter_rows(values_only=True)]
    wb.close()
    return data


def extract_system_prompt(rows: list, prompt_id: str = None) -> str:
    """
    从提示词 sheet 中提取 system prompt

    Args:
        rows: 提示词 sheet 的数据
        prompt_id: 提示词ID（如果提供，按ID查找；否则使用"是否训练"列）
    """
    header = [str(c).strip() if c else "" for c in rows[0]]

    # 优先按 prompt_id 查找（支持大小写不敏感和模糊匹配）
    if prompt_id:
        try:
            col_id = header.index("提示词ID")
            col_system = header.index("System 提示词")
            target_id = str(prompt_id).strip().lower()

            # 先尝试精确匹配
            for row in rows[1:]:
                if len(row) > max(col_id, col_system):
                    row_id = str(row[col_id]).strip().lower() if row[col_id] else ""
                    if row_id == target_id:
                        original_id = str(row[col_id]).strip()
                        prompt = str(row[col_system]).strip()
                        logger.info(f"使用指定提示词ID: {prompt_id} (精确匹配: {original_id})")
                        return prompt

            # 精确匹配失败，尝试模糊匹配（包含关系）
            matches = []
            for row in rows[1:]:
                if len(row) > max(col_id, col_system):
                    original_id = str(row[col_id]).strip() if row[col_id] else ""
                    row_id_lower = original_id.lower()
                    if target_id in row_id_lower:
                        prompt = str(row[col_system]).strip()
                        matches.append((original_id, row_id_lower, prompt))

            # 选择最短的匹配（最精确的）
            if matches:
                matches.sort(key=lambda x: len(x[1]))
                best_match_id, _, best_prompt = matches[0]
                logger.info(f"使用指定提示词ID: {prompt_id} (模糊匹配: {best_match_id})")
                return best_prompt

            logger.warning(f"未找到提示词ID '{prompt_id}'，回退到默认逻辑")
        except ValueError:
            logger.warning(f"未找到 '提示词ID' 列，回退到默认逻辑")

    # 默认逻辑：根据"是否训练"列查找
    try:
        col_system = header.index("System 提示词")
        col_train = header.index("是否训练")
        for row in rows[1:]:
            if len(row) > col_train and str(row[col_train]).strip() == "是":
                logger.info(f"使用默认提示词（是否训练=是）")
                return str(row[col_system]).strip()
    except ValueError:
        logger.warning(f"未找到必需的列：System 提示词 或 是否训练")

    return ""


def build_code_map(rows: list, prompt: str) -> tuple:
    """返回 (名称->编码映射, 小写索引, 编码->名称反查)"""
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_code = header.index("业务编码")
    col_name = header.index("业务名称")

    mapping = {}
    code_to_name = {}
    for row in rows[1:]:
        if not row or len(row) <= max(col_code, col_name):
            continue
        code, name = row[col_code], row[col_name]
        if code and name:
            code, name = str(code).strip(), str(name).strip()
            mapping[name] = code
            code_to_name[code] = name

    # 从提示词中补充
    for m in re.finditer(r'【(\d+)】(.+?)(?:[：:\n])', prompt):
        name, code = m.group(2).strip(), m.group(1)
        if name not in mapping:
            mapping[name] = code
            code_to_name.setdefault(code, name)

    # 大小写兼容
    lower_map = {k.lower(): v for k, v in mapping.items()}

    # 同义词
    synonyms = {"查附近网点": "查询服务网点"}
    for alias, target in synonyms.items():
        if alias not in mapping and target in mapping:
            mapping[alias] = mapping[target]
            lower_map[alias.lower()] = mapping[target]

    return mapping, lower_map, code_to_name


def convert_to_jsonl(rows: list, system_prompt: str, code_map: dict, lower_map: dict, use_mapping: bool = True) -> tuple:
    header = [str(c).strip() if c else "" for c in rows[0]]
    col_input = header.index("客户问题")
    col_output = header.index("人工标注结果")
    HEADER_VALUES = set(header)

    result = []
    skipped = 0
    unmapped_labels = {}
    output_dist = defaultdict(int)

    for row in rows[1:]:
        if not row or len(row) <= max(col_input, col_output):
            continue
        query, label = row[col_input], row[col_output]
        if not query or not label:
            continue
        query, label = str(query).strip(), str(label).strip()

        if query in HEADER_VALUES or label in HEADER_VALUES:
            skipped += 1
            continue

        if use_mapping:
            # 使用映射：将标签转换为编码
            code = code_map.get(label) or lower_map.get(label.lower())
            if not code:
                unmapped_labels[label] = unmapped_labels.get(label, 0) + 1
                skipped += 1
                continue
            output_value = code
        else:
            # 不使用映射：直接使用原始标签
            output_value = label

        result.append({"instruction": system_prompt, "input": query, "output": output_value})
        output_dist[output_value] += 1

    return result, skipped, unmapped_labels, output_dist


def save_jsonl(data: list, path: str):
    with open(path, "w", encoding="utf-8") as f:
        for item in data:
            f.write(json.dumps(item, ensure_ascii=False) + "\n")


def log_distribution(title: str, output_dist: dict, code_to_name: dict, unmapped: dict, skipped: int, use_mapping: bool = True):
    logger.info(f"{'='*50}  {title}  {'='*50}")
    if use_mapping:
        # 使用映射时，显示编码和名称
        for code in sorted(output_dist.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            name = code_to_name.get(code, "")
            logger.info(f"  [{name}-{code}] : {output_dist[code]} 条")
    else:
        # 不使用映射时，直接显示标签
        for label, count in sorted(output_dist.items(), key=lambda x: x[1], reverse=True):
            logger.info(f"  [{label}] : {count} 条")
    logger.info(f"  {'─'*40}")
    logger.info(f"  合计: {sum(output_dist.values())} 条")
    if unmapped:
        logger.warning(f"  未映射标签 (跳过 {skipped} 条):")
        for label, count in unmapped.items():
            logger.warning(f"    - '{label}': {count} 条")


def update_dataset_info(dataset_name: str, train_file: str, dataset_dir: str):
    """注册数据集到 dataset_info.json（仅训练集，验证集不参与训练）"""
    info_path = os.path.join(dataset_dir, "dataset_info.json")
    if os.path.exists(info_path):
        with open(info_path, "r", encoding="utf-8") as f:
            info = json.load(f)
    else:
        info = {}
    info[dataset_name] = {
        "file_name": train_file,
        "columns": {
            "prompt": "instruction",
            "query": "input",
            "response": "output",
        },
    }
    with open(info_path, "w", encoding="utf-8") as f:
        json.dump(info, f, ensure_ascii=False, indent=2)
    logger.info(f"已注册到 {os.path.join(dataset_dir, 'dataset_info.json')}")


def update_train_config(dataset_name: str):
    """自动更新 train_config.yaml 中的 dataset_name"""
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, "config", "train_config.yaml")
    if not os.path.exists(config_path):
        logger.warning(f"未找到配置文件: {config_path}")
        return

    with open(config_path, "r", encoding="utf-8") as f:
        content = f.read()

    new_content, count = re.subn(
        r'(dataset_name:\s*").*?(")',
        rf'\g<1>{dataset_name}\g<2>',
        content,
    )
    if count > 0:
        with open(config_path, "w", encoding="utf-8") as f:
            f.write(new_content)
        logger.info(f"已更新 config/train_config.yaml: dataset_name -> {dataset_name}")
    else:
        logger.warning(f"未在 train_config.yaml 中找到 dataset_name 字段")


def main():
    parser = argparse.ArgumentParser(description="从 Excel 生成微调训练数据")
    parser.add_argument("--input", nargs="?", default=None, help="Excel 文件路径 (默认选最新)")
    parser.add_argument("--output_dir", default="data", help="输出目录")
    parser.add_argument("--dataset_name", default="mainintent", help="数据集名称")
    parser.add_argument("--raw_output", action="store_true", help="不使用编码映射，直接使用人工标注结果作为 output")
    parser.add_argument("--prompt_id", default=None, help="提示词ID（指定使用哪个提示词，优先级高于默认逻辑）")
    args = parser.parse_args()

    use_mapping = not args.raw_output

    # 确定输入文件
    if args.input:
        excel_path = args.input
    else:
        excel_path = find_latest_excel()

    os.makedirs(args.output_dir, exist_ok=True)

    # 确定运行目录和日志路径
    from step2_train import generate_run_name, load_config
    config_path = "config/train_config.yaml"
    try:
        cfg = load_config(config_path)
        base_name = generate_run_name(cfg)
        timestamp = datetime.now().strftime("%m%d_%H%M")
        run_name = f"{base_name}_{timestamp}"

        # 从配置文件读取 prompt_id（如果命令行未指定）
        if args.prompt_id is None:
            args.prompt_id = cfg.get("data", {}).get("prompt_id", None)
    except Exception:
        run_name = f"prepare_data_{datetime.now().strftime('%m%d_%H%M')}"
    run_dir = os.path.join("outputs", run_name)
    log_ts = datetime.now().strftime("%m%d_%H%M")
    log_path = os.path.join(run_dir, "logs", f"step1_prepare_{log_ts}.log")

    global logger
    logger = setup_logger(log_path)

    logger.info(f"{'='*60}")
    logger.info(f"  Step 1 — 数据准备")
    logger.info(f"{'='*60}")
    logger.info(f"  运行目录 : {run_dir}")
    logger.info(f"  数据文件 : {os.path.basename(excel_path)}")
    logger.info(f"  日志文件 : {log_path}")

    sheets = read_excel(excel_path)

    # 提示词
    logger.info(f"{'─'*60}")
    logger.info(f"  系统提示词")
    logger.info(f"{'─'*60}")
    system_prompt = extract_system_prompt(sheets.get("提示词", []), args.prompt_id)
    logger.info(f"  长度: {len(system_prompt)} 字符")
    logger.info(f"  {'·'*40}")
    for line in system_prompt.split("\n"):
        logger.info(f"  {line}")
    logger.info(f"  {'·'*40}")

    # 编码映射
    logger.info(f"{'─'*60}")
    logger.info(f"  编码映射")
    logger.info(f"{'─'*60}")
    if use_mapping:
        code_map, lower_map, code_to_name = build_code_map(
            sheets.get("业务编码映射", []), system_prompt
        )
        logger.info(f"  映射条数: {len(code_map)}")
        for code in sorted(code_to_name.keys(), key=lambda x: int(x) if x.isdigit() else 0):
            name = code_to_name[code]
            logger.info(f"    [{name}] -> {code}")
    else:
        code_map, lower_map, code_to_name = {}, {}, {}
        logger.info(f"  输出模式: 原始标签（不使用编码映射）")

    # 训练集
    train_data, train_skip, train_unmapped, train_dist = convert_to_jsonl(
        sheets.get("训练集", []), system_prompt, code_map, lower_map, use_mapping
    )
    train_file = f"{args.dataset_name}_train.jsonl"
    train_path = os.path.join(args.output_dir, train_file)
    save_jsonl(train_data, train_path)
    log_distribution("训练集", train_dist, code_to_name, train_unmapped, train_skip, use_mapping)
    logger.info(f"  已保存: {train_path}")

    # 测试集（训练完成后用于推理测试，不参与训练）
    val_data, val_skip, val_unmapped, val_dist = convert_to_jsonl(
        sheets.get("测试集", []), system_prompt, code_map, lower_map, use_mapping
    )
    val_file = f"{args.dataset_name}_val.jsonl"
    val_path = os.path.join(args.output_dir, val_file)
    save_jsonl(val_data, val_path)
    log_distribution("测试集（训练后评估用）", val_dist, code_to_name, val_unmapped, val_skip, use_mapping)
    logger.info(f"  已保存: {val_path}")

    # 注册 & 更新配置
    logger.info(f"{'─'*60}")
    logger.info(f"  配置更新")
    logger.info(f"{'─'*60}")
    update_dataset_info(args.dataset_name, train_file, args.output_dir)
    update_train_config(args.dataset_name)

    logger.info(f"{'='*60}")
    logger.info(f"  数据准备完成")
    logger.info(f"{'='*60}")


if __name__ == "__main__":
    main()
